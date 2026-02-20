"""Audit Report Generator - Export comprehensive calculation documentation.

This module generates audit-ready exports showing all formulas, inputs,
and traced calculations for regulatory review and stakeholder transparency.
"""

import io
from datetime import datetime
from typing import Optional, Dict, List
from dataclasses import dataclass

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.calculations.detailed_cashflow import DetailedCashFlowResult
from src.calculations.trace import TraceContext, TracedValue
from src.calculations.formula_registry import FormulaRegistry, FormulaCategory


@dataclass
class AuditReportConfig:
    """Configuration for audit report generation."""
    include_formula_registry: bool = True
    include_traced_values: bool = True
    include_cash_flows: bool = True
    include_sources_uses: bool = True
    include_summary: bool = True
    project_name: str = "Real Estate Development"
    scenario_name: str = "Analysis"


def _format_value(value: float, unit: str = "$") -> str:
    """Format a value for display in reports."""
    if unit == "%":
        return f"{value:.2%}"
    elif abs(value) >= 1_000_000:
        return f"${value/1_000_000:,.2f}M"
    elif abs(value) >= 1_000:
        return f"${value/1_000:,.1f}K"
    elif value == 0:
        return "$0"
    else:
        return f"${value:,.0f}"


def _add_header_style(ws, row: int, cols: int) -> None:
    """Apply header styling to a row."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _add_section_header(ws, title: str, row: int) -> int:
    """Add a section header and return next row."""
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    return row + 1


def generate_audit_excel(
    result: DetailedCashFlowResult,
    config: Optional[AuditReportConfig] = None,
) -> bytes:
    """Generate a comprehensive Excel audit report.

    Args:
        result: The DetailedCashFlowResult from calculate_deal()
        config: Optional configuration for the report

    Returns:
        Excel file as bytes
    """
    if config is None:
        config = AuditReportConfig()

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # === Sheet 1: Summary ===
    if config.include_summary:
        ws = wb.create_sheet("Summary")
        _create_summary_sheet(ws, result, config)

    # === Sheet 2: Sources & Uses ===
    if config.include_sources_uses:
        ws = wb.create_sheet("Sources & Uses")
        _create_sources_uses_sheet(ws, result)

    # === Sheet 3: Formula Registry ===
    if config.include_formula_registry:
        ws = wb.create_sheet("Formula Registry")
        _create_formula_registry_sheet(ws)

    # === Sheet 4: Traced Calculations ===
    if config.include_traced_values and result.trace_context:
        ws = wb.create_sheet("Traced Calculations")
        _create_traced_calculations_sheet(ws, result.trace_context)

    # === Sheet 5: Cash Flows ===
    if config.include_cash_flows:
        ws = wb.create_sheet("Cash Flows")
        _create_cash_flows_sheet(ws, result)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _create_summary_sheet(
    ws,
    result: DetailedCashFlowResult,
    config: AuditReportConfig
) -> None:
    """Create the summary sheet."""
    row = 1

    # Title
    ws.cell(row=row, column=1, value=f"Audit Report: {config.project_name}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=16)
    row += 1

    ws.cell(row=row, column=1, value=f"Scenario: {config.scenario_name}")
    row += 1

    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    row += 2

    # Key Metrics
    row = _add_section_header(ws, "Key Metrics", row)
    row += 1

    metrics = [
        ("Total Development Cost", f"${result.sources_uses.tdc:,.0f}"),
        ("Equity Required", f"${result.sources_uses.equity:,.0f}"),
        ("Construction Loan", f"${result.sources_uses.construction_loan:,.0f}"),
        ("Permanent Loan", f"${result.perm_loan_amount:,.0f}"),
        ("", ""),
        ("Levered IRR", f"{result.levered_irr:.2%}"),
        ("Unlevered IRR", f"{result.unlevered_irr:.2%}"),
        ("Equity Multiple", f"{result.equity_multiple:.2f}x"),
        ("Yield on Cost", f"{result.yield_on_cost:.2%}"),
        ("", ""),
        ("Stabilized NOI", f"${result.stabilized_noi:,.0f}"),
        ("Reversion Value", f"${result.reversion_value:,.0f}"),
    ]

    for label, value in metrics:
        if label:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
        row += 1

    # Timeline
    row += 1
    row = _add_section_header(ws, "Timeline", row)
    row += 1

    timeline = [
        ("Predevelopment", f"{result.predevelopment_end} months"),
        ("Construction", f"{result.construction_end - result.predevelopment_end} months"),
        ("Lease-up", f"{result.leaseup_end - result.construction_end} months"),
        ("Operations", f"{result.operations_end - result.leaseup_end} months"),
        ("Total Periods", f"{result.total_periods} months"),
    ]

    for label, value in timeline:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def _create_sources_uses_sheet(ws, result: DetailedCashFlowResult) -> None:
    """Create the Sources & Uses sheet."""
    su = result.sources_uses

    row = 1
    row = _add_section_header(ws, "Uses of Funds", row)
    row += 1

    # Headers
    ws.cell(row=row, column=1, value="Item")
    ws.cell(row=row, column=2, value="Amount")
    ws.cell(row=row, column=3, value="% of TDC")
    ws.cell(row=row, column=4, value="Formula")
    _add_header_style(ws, row, 4)
    row += 1

    uses = [
        ("Land", su.land, "Input"),
        ("Hard Costs", su.hard_costs, "units x cost_per_unit x (1 + contingency)"),
        ("Soft Costs", su.soft_costs, "hard_costs x soft_cost_pct"),
        ("Interest During Construction", su.idc, "Sum of monthly interest on construction loan"),
        ("", "", ""),
        ("Total Development Cost", su.tdc, "land + hard + soft + idc"),
    ]

    for label, amount, formula in uses:
        if label:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=f"${amount:,.0f}")
            ws.cell(row=row, column=3, value=f"{amount/su.tdc:.1%}" if su.tdc > 0 else "-")
            ws.cell(row=row, column=4, value=formula)
            if label == "Total Development Cost":
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2).font = Font(bold=True)
        row += 1

    row += 1
    row = _add_section_header(ws, "Sources of Funds", row)
    row += 1

    # Headers
    ws.cell(row=row, column=1, value="Item")
    ws.cell(row=row, column=2, value="Amount")
    ws.cell(row=row, column=3, value="% of TDC")
    ws.cell(row=row, column=4, value="Formula")
    _add_header_style(ws, row, 4)
    row += 1

    sources = [
        ("Developer Equity", su.equity, "tdc x (1 - ltc)"),
        ("Construction Loan", su.construction_loan, "tdc x ltc"),
        ("", "", ""),
        ("Total Sources", su.total_sources, "equity + construction_loan"),
    ]

    for label, amount, formula in sources:
        if label:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=f"${amount:,.0f}")
            ws.cell(row=row, column=3, value=f"{amount/su.tdc:.1%}" if su.tdc > 0 else "-")
            ws.cell(row=row, column=4, value=formula)
            if label == "Total Sources":
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2).font = Font(bold=True)
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45


def _create_formula_registry_sheet(ws) -> None:
    """Create the Formula Registry sheet."""
    all_formulas = FormulaRegistry.get_all()

    row = 1
    row = _add_section_header(ws, "Formula Registry - All Calculation Definitions", row)
    row += 2

    # Headers
    headers = ["Category", "Name", "Field Path", "Formula", "Inputs", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _add_header_style(ws, row, len(headers))
    row += 1

    # Group by category
    by_category: Dict[FormulaCategory, list] = {}
    for field_path, formula in all_formulas.items():
        if formula.category not in by_category:
            by_category[formula.category] = []
        by_category[formula.category].append((field_path, formula))

    for category in FormulaCategory:
        if category not in by_category:
            continue

        for field_path, formula in sorted(by_category[category], key=lambda x: x[0]):
            ws.cell(row=row, column=1, value=category.value)
            ws.cell(row=row, column=2, value=formula.name)
            ws.cell(row=row, column=3, value=field_path)
            ws.cell(row=row, column=4, value=formula.formula)
            ws.cell(row=row, column=5, value=", ".join(formula.inputs) if formula.inputs else "-")
            ws.cell(row=row, column=6, value=formula.notes if formula.notes else "-")
            row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40


def _create_traced_calculations_sheet(ws, trace_context: TraceContext) -> None:
    """Create the Traced Calculations sheet."""
    row = 1
    row = _add_section_header(ws, "Traced Calculations - Actual Values Used", row)
    row += 2

    # Headers
    headers = ["Field Path", "Result", "Computed Formula", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _add_header_style(ws, row, len(headers))
    row += 1

    # Sort traces by field path
    for trace_key in sorted(trace_context.traces.keys()):
        trace = trace_context.traces[trace_key]

        ws.cell(row=row, column=1, value=trace.field_path)
        ws.cell(row=row, column=2, value=_format_value(trace.value))
        ws.cell(row=row, column=3, value=trace.computed_formula[:100] if len(trace.computed_formula) > 100 else trace.computed_formula)
        ws.cell(row=row, column=4, value=trace.notes if trace.notes else "-")
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 30


def _create_cash_flows_sheet(ws, result: DetailedCashFlowResult) -> None:
    """Create the Cash Flows sheet."""
    row = 1
    row = _add_section_header(ws, "Period-by-Period Cash Flows", row)
    row += 2

    # Headers
    headers = ["Period", "Phase", "GPR", "Vacancy", "EGI", "OpEx", "NOI", "Debt Service", "Levered CF"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    _add_header_style(ws, row, len(headers))
    row += 1

    # Data rows
    for period in result.periods:
        h = period.header
        ops = period.operations

        phase = "REVERSION" if h.is_reversion else (
            "OPS" if h.is_operations else (
                "LEASEUP" if h.is_leaseup else (
                    "CONSTR" if h.is_construction else "PREDEV"
                )
            )
        )

        ws.cell(row=row, column=1, value=h.period)
        ws.cell(row=row, column=2, value=phase)
        ws.cell(row=row, column=3, value=f"${ops.gpr:,.0f}")
        ws.cell(row=row, column=4, value=f"${ops.less_vacancy:,.0f}")
        ws.cell(row=row, column=5, value=f"${ops.egi:,.0f}")
        ws.cell(row=row, column=6, value=f"${ops.less_opex_ex_taxes + ops.less_property_taxes:,.0f}")
        ws.cell(row=row, column=7, value=f"${ops.noi:,.0f}")
        ws.cell(row=row, column=8, value=f"${period.net_debt_cf:,.0f}")
        ws.cell(row=row, column=9, value=f"${period.levered_cf:,.0f}")
        row += 1

    # Adjust column widths
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 15
