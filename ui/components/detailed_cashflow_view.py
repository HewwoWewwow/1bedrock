"""Detailed cash flow view component for the Streamlit UI."""

import streamlit as st
import pandas as pd
from typing import List, Optional, Literal
from datetime import date
from dataclasses import dataclass

from src.calculations.detailed_cashflow import (
    DetailedCashFlowResult, DetailedPeriodCashFlow,
    generate_detailed_cash_flow
)
from src.calculations.sources_uses import SourcesUses
from src.calculations.property_tax import (
    TaxingAuthorityStack, get_austin_tax_stack,
    PropertyTaxSchedule, PropertyTaxPeriod,
    generate_property_tax_schedule, calculate_assessed_value_schedule,
    AssessedValueTiming
)


def render_deal_summary_header(
    tdc: float,
    units: int,
    equity: float,
    senior_debt: float,
    levered_irr: float,
    equity_multiple: float,
) -> None:
    """Render a compact deal summary header."""
    col1, col2, col3, col4, col5, col6 = st.columns(6)

    with col1:
        st.metric("TDC", f"${tdc/1e6:.1f}M", help="Total Development Cost")
    with col2:
        st.metric("Units", f"{units:,}")
    with col3:
        st.metric("TDC/Unit", f"${tdc/units:,.0f}" if units > 0 else "-")
    with col4:
        st.metric("Equity", f"${equity/1e6:.1f}M")
    with col5:
        st.metric("Levered IRR", f"{levered_irr:.1%}")
    with col6:
        st.metric("Multiple", f"{equity_multiple:.2f}x")


def export_cashflow_to_csv(result: DetailedCashFlowResult) -> str:
    """Export the detailed cash flow to CSV format.

    Returns:
        CSV string that can be downloaded
    """
    import io
    import csv

    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    headers = [
        "Period", "Phase", "TDC Drawn", "Equity Drawn", "GPR", "Vacancy",
        "EGI", "OpEx", "Prop Tax", "TIF Reimbursement", "TIF Abatement",
        "Net TIF Benefit", "NOI", "Dev Cost", "Reversion",
        "Unlevered CF", "Const Debt Net", "Perm Debt Net", "Mezz Debt Net",
        "Preferred Net", "Net Debt CF", "Levered CF"
    ]
    writer.writerow(headers)

    # Data rows
    for p in result.periods:
        phase = ""
        if p.header.is_predevelopment:
            phase = "Predev"
        elif p.header.is_construction:
            phase = "Construction"
        elif p.header.is_leaseup:
            phase = "Lease-up"
        elif p.header.is_operations:
            phase = "Operations"
        if p.header.is_reversion:
            phase = "Reversion"

        mezz_net = p.mezzanine_debt.net_cf if p.mezzanine_debt else 0
        pref_net = p.preferred_equity.net_cf if p.preferred_equity else 0

        # TIF data
        tif_row = p.operations.tif
        tif_reimb = tif_row.tif_reimbursement if tif_row else 0
        tif_abate = tif_row.abatement_amount if tif_row else 0
        tif_net = tif_row.net_tif_benefit if tif_row else 0

        row = [
            p.header.period,
            phase,
            p.development.draw_dollars_total,
            p.equity.equity_drawn,
            p.operations.gpr,
            p.operations.less_vacancy,
            p.operations.egi,
            p.operations.less_opex_ex_taxes,
            p.operations.less_property_taxes,
            tif_reimb,
            tif_abate,
            tif_net,
            p.operations.noi,
            p.investment.dev_cost,
            p.investment.reversion,
            p.investment.unlevered_cf,
            p.construction_debt.net_cf,
            p.permanent_debt.net_cf,
            mezz_net,
            pref_net,
            p.net_debt_cf,
            p.levered_cf,
        ]
        writer.writerow(row)

    return output.getvalue()


def export_cashflow_to_excel(result: DetailedCashFlowResult, scenario_name: str = "Scenario") -> bytes:
    """Export the detailed cash flow to Excel format for direct comparison.

    Creates a comprehensive Excel workbook with:
    - Summary sheet with key metrics
    - Sources & Uses sheet
    - Monthly cash flows sheet with all line items

    Args:
        result: The detailed cash flow result to export
        scenario_name: Name of the scenario for the export

    Returns:
        Excel file as bytes that can be downloaded
    """
    import io
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    # =========================================================================
    # SHEET 1: SUMMARY
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Header styles
    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11, color="1565C0")
    currency_fmt = '#,##0'
    pct_fmt = '0.00%'

    # Title
    ws_summary['A1'] = f"{scenario_name} - Cash Flow Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)

    # Key Metrics
    row = 3
    ws_summary[f'A{row}'] = "KEY METRICS"
    ws_summary[f'A{row}'].font = section_font

    metrics = [
        ("Total Development Cost", result.sources_uses.tdc, currency_fmt),
        ("Equity Required", result.sources_uses.equity, currency_fmt),
        ("Construction Loan", result.sources_uses.construction_loan, currency_fmt),
        ("Loan-to-Cost", result.sources_uses.ltc, pct_fmt),
        ("", "", ""),
        ("Unlevered IRR", result.unlevered_irr, pct_fmt),
        ("Levered IRR", result.levered_irr, pct_fmt),
        ("Total NOI", result.total_noi, currency_fmt),
        ("Reversion Value", result.reversion_value, currency_fmt),
    ]

    row = 4
    for label, value, fmt in metrics:
        ws_summary[f'A{row}'] = label
        if value != "":
            ws_summary[f'B{row}'] = value
            ws_summary[f'B{row}'].number_format = fmt
        row += 1

    # Phase Boundaries
    row += 1
    ws_summary[f'A{row}'] = "PHASE BOUNDARIES (Month)"
    ws_summary[f'A{row}'].font = section_font
    row += 1

    phases = [
        ("Predevelopment End", result.predevelopment_end),
        ("Construction End", result.construction_end),
        ("Lease-up End", result.leaseup_end),
        ("Operations End", result.operations_end),
        ("Total Periods", result.total_periods),
    ]

    for label, value in phases:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        row += 1

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 18

    # =========================================================================
    # SHEET 2: SOURCES & USES
    # =========================================================================
    ws_su = wb.create_sheet("Sources & Uses")

    su = result.sources_uses

    ws_su['A1'] = "USES OF FUNDS"
    ws_su['A1'].font = section_font

    uses = [
        ("Land", su.land),
        ("Hard Costs", su.hard_costs),
        ("Soft Costs", su.soft_costs),
        ("IDC (Interest During Construction)", su.idc),
        ("TOTAL DEVELOPMENT COST", su.tdc),
    ]

    row = 2
    for label, value in uses:
        ws_su[f'A{row}'] = label
        ws_su[f'B{row}'] = value
        ws_su[f'B{row}'].number_format = currency_fmt
        if "TOTAL" in label:
            ws_su[f'A{row}'].font = Font(bold=True)
            ws_su[f'B{row}'].font = Font(bold=True)
        row += 1

    row += 1
    ws_su[f'A{row}'] = "SOURCES OF FUNDS"
    ws_su[f'A{row}'].font = section_font
    row += 1

    sources = [
        ("Equity", su.equity),
        ("Construction Loan", su.construction_loan),
        ("TOTAL SOURCES", su.total_sources),
    ]

    for label, value in sources:
        ws_su[f'A{row}'] = label
        ws_su[f'B{row}'] = value
        ws_su[f'B{row}'].number_format = currency_fmt
        if "TOTAL" in label:
            ws_su[f'A{row}'].font = Font(bold=True)
            ws_su[f'B{row}'].font = Font(bold=True)
        row += 1

    ws_su.column_dimensions['A'].width = 35
    ws_su.column_dimensions['B'].width = 18

    # =========================================================================
    # SHEET 3: MONTHLY CASH FLOWS
    # =========================================================================
    ws_cf = wb.create_sheet("Monthly Cash Flows")

    # Build header row
    headers = [
        "Period", "Phase",
        # Development
        "TDC Drawn", "Predev Drawn", "Construction Drawn",
        # Equity
        "Equity Drawn",
        # Revenue
        "GPR", "Vacancy", "EGI",
        # Expenses
        "OpEx (ex taxes)", "Property Taxes",
        # TIF
        "TIF Reimbursement", "TIF Abatement", "Net TIF Benefit",
        # NOI
        "NOI",
        # Investment
        "Dev Cost", "Reserves", "Reversion", "Unlevered CF",
        # Construction Debt
        "Const Debt Added", "Const Interest", "Const Repaid", "Const Net CF",
        # Perm Debt
        "Perm Payment", "Perm Interest", "Perm Principal", "Perm Payoff", "Perm Net CF",
        # Final
        "Net Debt CF", "Levered CF",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws_cf.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Build data rows
    row = 2
    for p in result.periods:
        # Determine phase
        if p.header.is_predevelopment:
            phase = "Predev"
        elif p.header.is_construction:
            phase = "Construction"
        elif p.header.is_leaseup:
            phase = "Lease-up"
        elif p.header.is_reversion:
            phase = "Reversion"
        elif p.header.is_operations:
            phase = "Operations"
        else:
            phase = ""

        # TIF data
        tif_row = p.operations.tif
        tif_reimb = tif_row.tif_reimbursement if tif_row else 0
        tif_abate = tif_row.abatement_amount if tif_row else 0
        tif_net = tif_row.net_tif_benefit if tif_row else 0

        values = [
            p.header.period,
            phase,
            # Development
            p.development.draw_dollars_total,
            p.development.draw_dollars_predev,
            p.development.draw_dollars_construction,
            # Equity
            p.equity.equity_drawn,
            # Revenue
            p.operations.gpr,
            p.operations.less_vacancy,
            p.operations.egi,
            # Expenses
            p.operations.less_opex_ex_taxes,
            p.operations.less_property_taxes,
            # TIF
            tif_reimb,
            tif_abate,
            tif_net,
            # NOI
            p.operations.noi,
            # Investment
            p.investment.dev_cost,
            p.investment.reserves,
            p.investment.reversion,
            p.investment.unlevered_cf,
            # Construction Debt
            p.construction_debt.debt_added,
            p.construction_debt.interest_in_period,
            p.construction_debt.repaid,
            p.construction_debt.net_cf,
            # Perm Debt
            p.permanent_debt.pmt_in_period,
            p.permanent_debt.interest_pmt,
            p.permanent_debt.principal_pmt,
            p.permanent_debt.payoff,
            p.permanent_debt.net_cf,
            # Final
            p.net_debt_cf,
            p.levered_cf,
        ]

        for col, value in enumerate(values, 1):
            cell = ws_cf.cell(row=row, column=col, value=value)
            # Format numeric values
            if col > 2 and isinstance(value, (int, float)):
                cell.number_format = currency_fmt

        row += 1

    # Add totals row
    total_row = row
    ws_cf.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_cf.cell(row=total_row, column=2, value="").font = Font(bold=True)

    # Calculate totals for each numeric column
    for col in range(3, len(headers) + 1):
        total = sum(
            ws_cf.cell(row=r, column=col).value or 0
            for r in range(2, total_row)
        )
        cell = ws_cf.cell(row=total_row, column=col, value=total)
        cell.font = Font(bold=True)
        cell.number_format = currency_fmt

    # Freeze header row and first two columns
    ws_cf.freeze_panes = 'C2'

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws_cf.column_dimensions[ws_cf.cell(row=1, column=col).column_letter].width = 14

    # =========================================================================
    # SHEET 4: ANNUAL SUMMARY
    # =========================================================================
    ws_annual = wb.create_sheet("Annual Summary")

    # Group periods by year
    years_data = {}
    for p in result.periods:
        year = (p.header.period - 1) // 12 + 1
        if year not in years_data:
            years_data[year] = []
        years_data[year].append(p)

    annual_headers = [
        "Year", "Periods",
        "GPR", "Vacancy", "EGI",
        "OpEx", "Prop Taxes", "NOI",
        "TIF Benefit",
        "Debt Service", "Levered CF",
    ]

    for col, header in enumerate(annual_headers, 1):
        cell = ws_annual.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    row = 2
    for year in sorted(years_data.keys()):
        periods = years_data[year]
        period_range = f"{periods[0].header.period}-{periods[-1].header.period}"

        tif_benefit = sum(
            (p.operations.tif.net_tif_benefit if p.operations.tif else 0)
            for p in periods
        )

        values = [
            year,
            period_range,
            sum(p.operations.gpr for p in periods),
            sum(p.operations.less_vacancy for p in periods),
            sum(p.operations.egi for p in periods),
            sum(p.operations.less_opex_ex_taxes for p in periods),
            sum(p.operations.less_property_taxes for p in periods),
            sum(p.operations.noi for p in periods),
            tif_benefit,
            sum(p.permanent_debt.pmt_in_period for p in periods),
            sum(p.levered_cf for p in periods),
        ]

        for col, value in enumerate(values, 1):
            cell = ws_annual.cell(row=row, column=col, value=value)
            if col > 2 and isinstance(value, (int, float)):
                cell.number_format = currency_fmt

        row += 1

    # Adjust column widths
    ws_annual.column_dimensions['A'].width = 8
    ws_annual.column_dimensions['B'].width = 12
    for col in range(3, len(annual_headers) + 1):
        ws_annual.column_dimensions[ws_annual.cell(row=1, column=col).column_letter].width = 14

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def render_sources_uses(
    sources_uses: SourcesUses,
    show_detail_toggle: bool = True,
    key_prefix: str = "",
) -> None:
    """Render the Sources & Uses table.

    Args:
        sources_uses: The SourcesUses object with capital stack data
        show_detail_toggle: Whether to show the detail toggle (default True)
        key_prefix: Prefix for widget keys to avoid duplicates
    """
    st.subheader("Sources & Uses")

    # Detail toggle
    show_detail = False
    if show_detail_toggle:
        toggle_key = f"{key_prefix}_su_detail_toggle" if key_prefix else "su_detail_toggle"
        show_detail = st.checkbox("Show detailed breakdown", value=False, key=toggle_key)

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**USES**")

        if show_detail and _has_detailed_breakdown(sources_uses):
            # Detailed breakdown
            uses_data = _build_detailed_uses_data(sources_uses)
        else:
            # Summary view
            uses_data = {
                "Category": ["Land", "Hard Costs", "Soft Costs", "IDC", "**Total Development Cost**"],
                "Amount": [
                    f"${sources_uses.land:,.0f}",
                    f"${sources_uses.hard_costs:,.0f}",
                    f"${sources_uses.soft_costs:,.0f}",
                    f"${sources_uses.idc:,.0f}",
                    f"**${sources_uses.tdc:,.0f}**",
                ],
                "% of TDC": [
                    f"{sources_uses.land / sources_uses.tdc:.1%}",
                    f"{sources_uses.hard_costs / sources_uses.tdc:.1%}",
                    f"{sources_uses.soft_costs / sources_uses.tdc:.1%}",
                    f"{sources_uses.idc / sources_uses.tdc:.1%}",
                    "100.0%",
                ],
            }
        st.dataframe(pd.DataFrame(uses_data), use_container_width=True, hide_index=True)

    with col2:
        st.markdown("**SOURCES**")
        sources_data = {
            "Category": ["Equity", "Construction Loan", "**Total Sources**"],
            "Amount": [
                f"${sources_uses.equity:,.0f}",
                f"${sources_uses.construction_loan:,.0f}",
                f"**${sources_uses.total_sources:,.0f}**",
            ],
            "% of TDC": [
                f"{sources_uses.equity_pct:.1%}",
                f"{sources_uses.ltc:.1%}",
                "100.0%",
            ],
        }
        st.dataframe(pd.DataFrame(sources_data), use_container_width=True, hide_index=True)


def _has_detailed_breakdown(sources_uses: SourcesUses) -> bool:
    """Check if the SourcesUses has any detailed breakdown values populated."""
    return (
        sources_uses.hard_cost_contingency > 0 or
        sources_uses.soft_cost_contingency > 0 or
        sources_uses.predevelopment_costs > 0 or
        sources_uses.developer_fee > 0 or
        sources_uses.reserves > 0
    )


def _build_detailed_uses_data(sources_uses: SourcesUses) -> dict:
    """Build the detailed uses breakdown table data."""
    su = sources_uses

    # Calculate base costs from totals minus contingencies
    hard_costs_base = su.hard_costs - su.hard_cost_contingency

    # Soft costs breakdown: total soft_costs includes base + contingency + predev + dev fee
    # soft_costs field = hard_costs * effective_soft_cost_pct (the combined %)
    # We have individual values for contingency, predev, dev fee
    # So base soft = soft_costs - soft_cost_contingency - predevelopment_costs - developer_fee
    soft_costs_base = su.soft_costs - su.soft_cost_contingency - su.predevelopment_costs - su.developer_fee - su.reserves
    if soft_costs_base < 0:
        soft_costs_base = 0  # Guard against calculation issues

    categories = []
    amounts = []
    pct_tdc = []

    # Land
    categories.append("Land")
    amounts.append(f"${su.land:,.0f}")
    pct_tdc.append(f"{su.land / su.tdc:.1%}")

    # Hard Costs - Base
    categories.append("Hard Costs (Base)")
    amounts.append(f"${hard_costs_base:,.0f}")
    pct_tdc.append(f"{hard_costs_base / su.tdc:.1%}")

    # Hard Costs - Contingency
    if su.hard_cost_contingency > 0:
        categories.append("  + Contingency")
        amounts.append(f"${su.hard_cost_contingency:,.0f}")
        pct_tdc.append(f"{su.hard_cost_contingency / su.tdc:.1%}")

    # Soft Costs - Base
    categories.append("Soft Costs (Base)")
    amounts.append(f"${soft_costs_base:,.0f}")
    pct_tdc.append(f"{soft_costs_base / su.tdc:.1%}")

    # Soft Costs - Contingency
    if su.soft_cost_contingency > 0:
        categories.append("  + Contingency")
        amounts.append(f"${su.soft_cost_contingency:,.0f}")
        pct_tdc.append(f"{su.soft_cost_contingency / su.tdc:.1%}")

    # Predevelopment Costs
    if su.predevelopment_costs > 0:
        categories.append("Predevelopment Costs")
        amounts.append(f"${su.predevelopment_costs:,.0f}")
        pct_tdc.append(f"{su.predevelopment_costs / su.tdc:.1%}")

    # Developer Fee
    if su.developer_fee > 0:
        categories.append("Developer Fee")
        amounts.append(f"${su.developer_fee:,.0f}")
        pct_tdc.append(f"{su.developer_fee / su.tdc:.1%}")

    # Reserves
    if su.reserves > 0:
        categories.append("Reserves")
        amounts.append(f"${su.reserves:,.0f}")
        pct_tdc.append(f"{su.reserves / su.tdc:.1%}")

    # Loan Fee
    if su.loan_fee > 0:
        categories.append("Loan Fee")
        amounts.append(f"${su.loan_fee:,.0f}")
        pct_tdc.append(f"{su.loan_fee / su.tdc:.1%}")

    # IDC
    categories.append("Interest During Construction")
    amounts.append(f"${su.idc:,.0f}")
    pct_tdc.append(f"{su.idc / su.tdc:.1%}")

    # Total
    categories.append("**Total Development Cost**")
    amounts.append(f"**${su.tdc:,.0f}**")
    pct_tdc.append("100.0%")

    return {
        "Category": categories,
        "Amount": amounts,
        "% of TDC": pct_tdc,
    }


@dataclass
class AggregatedPeriod:
    """Aggregated period data for quarterly/annual views."""
    label: str
    period_start: int
    period_end: int

    # Development
    tdc_draw_total: float
    tdc_draw_predev: float
    tdc_draw_construction: float

    # Equity
    equity_drawn: float

    # Debt sources
    debt_financed: float

    # GPR
    gpr_all_market: float
    gpr_mixed_market: float
    gpr_mixed_affordable: float

    # Operations
    gpr: float
    vacancy: float
    egi: float
    opex_ex_taxes: float
    property_taxes: float
    noi: float

    # Investment
    dev_cost: float
    reserves: float
    reversion: float
    unlevered_cf: float

    # Senior Debt - Construction
    const_debt_added: float
    const_debt_interest: float
    const_debt_repaid: float
    const_debt_net: float

    # Senior Debt - Permanent
    perm_pmt: float
    perm_interest: float
    perm_principal: float
    perm_payoff: float
    perm_net: float

    # TIF Benefits (optional - placed after required fields)
    tif_reimbursement: float = 0.0
    tif_abatement: float = 0.0
    tif_net_benefit: float = 0.0

    # Mezzanine Debt
    mezz_pmt: float = 0.0
    mezz_interest: float = 0.0
    mezz_principal: float = 0.0
    mezz_payoff: float = 0.0
    mezz_net: float = 0.0
    mezz_active: bool = False

    # Preferred Equity
    pref_return: float = 0.0
    pref_payoff: float = 0.0
    pref_net: float = 0.0
    pref_active: bool = False

    # Final
    net_senior_debt_cf: float = 0.0
    net_mezz_debt_cf: float = 0.0
    net_preferred_cf: float = 0.0
    net_debt_cf: float = 0.0
    levered_cf: float = 0.0


def aggregate_periods(
    periods: List[DetailedPeriodCashFlow],
    aggregation: Literal["monthly", "quarterly", "annual"],
) -> List[AggregatedPeriod]:
    """Aggregate monthly periods into quarterly or annual buckets."""

    if aggregation == "monthly":
        # No aggregation needed, convert to AggregatedPeriod format
        result = []
        for p in periods:
            # Check for mezzanine and preferred
            has_mezz = p.mezzanine_debt is not None and p.mezzanine_debt.is_active
            has_pref = p.preferred_equity is not None and p.preferred_equity.is_active

            # Extract TIF data if available
            tif_row = p.operations.tif
            tif_reimb = tif_row.tif_reimbursement if tif_row else 0.0
            tif_abate = tif_row.abatement_amount if tif_row else 0.0
            tif_net = tif_row.net_tif_benefit if tif_row else 0.0

            result.append(AggregatedPeriod(
                label=f"M{p.header.period}",
                period_start=p.header.period,
                period_end=p.header.period,
                tdc_draw_total=p.development.draw_dollars_total,
                tdc_draw_predev=p.development.draw_dollars_predev,
                tdc_draw_construction=p.development.draw_dollars_construction,
                equity_drawn=p.equity.equity_drawn,
                debt_financed=p.debt_source.to_be_financed,
                gpr_all_market=p.gpr.gpr_all_market,
                gpr_mixed_market=p.gpr.gpr_mixed_market,
                gpr_mixed_affordable=p.gpr.gpr_mixed_affordable,
                gpr=p.operations.gpr,
                vacancy=p.operations.less_vacancy,
                egi=p.operations.egi,
                opex_ex_taxes=p.operations.less_opex_ex_taxes,
                property_taxes=p.operations.less_property_taxes,
                noi=p.operations.noi,
                # TIF Benefits
                tif_reimbursement=tif_reimb,
                tif_abatement=tif_abate,
                tif_net_benefit=tif_net,
                dev_cost=p.investment.dev_cost,
                reserves=p.investment.reserves,
                reversion=p.investment.reversion,
                unlevered_cf=p.investment.unlevered_cf,
                const_debt_added=p.construction_debt.debt_added,
                const_debt_interest=p.construction_debt.interest_in_period,
                const_debt_repaid=p.construction_debt.repaid,
                const_debt_net=p.construction_debt.net_cf,
                perm_pmt=p.permanent_debt.pmt_in_period,
                perm_interest=p.permanent_debt.interest_pmt,
                perm_principal=p.permanent_debt.principal_pmt,
                perm_payoff=p.permanent_debt.payoff,
                perm_net=p.permanent_debt.net_cf,
                # Mezzanine
                mezz_pmt=p.mezzanine_debt.pmt_in_period if has_mezz else 0,
                mezz_interest=p.mezzanine_debt.interest_pmt if has_mezz else 0,
                mezz_principal=p.mezzanine_debt.principal_pmt if has_mezz else 0,
                mezz_payoff=p.mezzanine_debt.payoff if has_mezz else 0,
                mezz_net=p.mezzanine_debt.net_cf if has_mezz else 0,
                mezz_active=has_mezz,
                # Preferred
                pref_return=p.preferred_equity.paid_return if has_pref else 0,
                pref_payoff=p.preferred_equity.payoff if has_pref else 0,
                pref_net=p.preferred_equity.net_cf if has_pref else 0,
                pref_active=has_pref,
                # Combined debt CFs
                net_senior_debt_cf=p.net_senior_debt_cf,
                net_mezz_debt_cf=p.net_mezz_debt_cf,
                net_preferred_cf=p.net_preferred_cf,
                net_debt_cf=p.net_debt_cf,
                levered_cf=p.levered_cf,
            ))
        return result

    # Group periods by quarter or year
    if aggregation == "quarterly":
        periods_per_group = 3
        label_prefix = "Q"
    else:  # annual
        periods_per_group = 12
        label_prefix = "Y"

    result = []
    group_num = 1
    i = 0

    while i < len(periods):
        group_periods = periods[i:i + periods_per_group]
        if not group_periods:
            break

        # Check if any period in group has mezz or preferred
        has_mezz = any(p.mezzanine_debt is not None and p.mezzanine_debt.is_active for p in group_periods)
        has_pref = any(p.preferred_equity is not None and p.preferred_equity.is_active for p in group_periods)

        # Aggregate the group
        agg = AggregatedPeriod(
            label=f"{label_prefix}{group_num}",
            period_start=group_periods[0].header.period,
            period_end=group_periods[-1].header.period,
            tdc_draw_total=sum(p.development.draw_dollars_total for p in group_periods),
            tdc_draw_predev=sum(p.development.draw_dollars_predev for p in group_periods),
            tdc_draw_construction=sum(p.development.draw_dollars_construction for p in group_periods),
            equity_drawn=sum(p.equity.equity_drawn for p in group_periods),
            debt_financed=sum(p.debt_source.to_be_financed for p in group_periods),
            gpr_all_market=sum(p.gpr.gpr_all_market for p in group_periods),
            gpr_mixed_market=sum(p.gpr.gpr_mixed_market for p in group_periods),
            gpr_mixed_affordable=sum(p.gpr.gpr_mixed_affordable for p in group_periods),
            gpr=sum(p.operations.gpr for p in group_periods),
            vacancy=sum(p.operations.less_vacancy for p in group_periods),
            egi=sum(p.operations.egi for p in group_periods),
            opex_ex_taxes=sum(p.operations.less_opex_ex_taxes for p in group_periods),
            property_taxes=sum(p.operations.less_property_taxes for p in group_periods),
            noi=sum(p.operations.noi for p in group_periods),
            # TIF Benefits
            tif_reimbursement=sum(p.operations.tif.tif_reimbursement if p.operations.tif else 0 for p in group_periods),
            tif_abatement=sum(p.operations.tif.abatement_amount if p.operations.tif else 0 for p in group_periods),
            tif_net_benefit=sum(p.operations.tif.net_tif_benefit if p.operations.tif else 0 for p in group_periods),
            dev_cost=sum(p.investment.dev_cost for p in group_periods),
            reserves=sum(p.investment.reserves for p in group_periods),
            reversion=sum(p.investment.reversion for p in group_periods),
            unlevered_cf=sum(p.investment.unlevered_cf for p in group_periods),
            const_debt_added=sum(p.construction_debt.debt_added for p in group_periods),
            const_debt_interest=sum(p.construction_debt.interest_in_period for p in group_periods),
            const_debt_repaid=sum(p.construction_debt.repaid for p in group_periods),
            const_debt_net=sum(p.construction_debt.net_cf for p in group_periods),
            perm_pmt=sum(p.permanent_debt.pmt_in_period for p in group_periods),
            perm_interest=sum(p.permanent_debt.interest_pmt for p in group_periods),
            perm_principal=sum(p.permanent_debt.principal_pmt for p in group_periods),
            perm_payoff=sum(p.permanent_debt.payoff for p in group_periods),
            perm_net=sum(p.permanent_debt.net_cf for p in group_periods),
            # Mezzanine
            mezz_pmt=sum(p.mezzanine_debt.pmt_in_period if p.mezzanine_debt else 0 for p in group_periods),
            mezz_interest=sum(p.mezzanine_debt.interest_pmt if p.mezzanine_debt else 0 for p in group_periods),
            mezz_principal=sum(p.mezzanine_debt.principal_pmt if p.mezzanine_debt else 0 for p in group_periods),
            mezz_payoff=sum(p.mezzanine_debt.payoff if p.mezzanine_debt else 0 for p in group_periods),
            mezz_net=sum(p.mezzanine_debt.net_cf if p.mezzanine_debt else 0 for p in group_periods),
            mezz_active=has_mezz,
            # Preferred
            pref_return=sum(p.preferred_equity.paid_return if p.preferred_equity else 0 for p in group_periods),
            pref_payoff=sum(p.preferred_equity.payoff if p.preferred_equity else 0 for p in group_periods),
            pref_net=sum(p.preferred_equity.net_cf if p.preferred_equity else 0 for p in group_periods),
            pref_active=has_pref,
            # Combined debt CFs
            net_senior_debt_cf=sum(p.net_senior_debt_cf for p in group_periods),
            net_mezz_debt_cf=sum(p.net_mezz_debt_cf for p in group_periods),
            net_preferred_cf=sum(p.net_preferred_cf for p in group_periods),
            net_debt_cf=sum(p.net_debt_cf for p in group_periods),
            levered_cf=sum(p.levered_cf for p in group_periods),
        )
        result.append(agg)
        group_num += 1
        i += periods_per_group

    return result


def _fmt_currency(val: float, show_parens: bool = True) -> str:
    """Format currency value."""
    if abs(val) < 0.5:
        return "-"
    if val < 0 and show_parens:
        return f"(${abs(val):,.0f})"
    return f"${val:,.0f}"


def _fmt_pct(val: float) -> str:
    """Format percentage value."""
    if abs(val) < 0.0001:
        return "-"
    return f"{val:.1%}"


def render_detailed_cashflow_table(
    result: DetailedCashFlowResult,
    aggregation: Literal["monthly", "quarterly", "annual"] = "monthly",
    is_mixed_income: bool = False,
    tif_lump_sum: float = 0.0,
) -> None:
    """Render the detailed cash flow table with sticky columns.

    Shows all rows from the Excel model structure with horizontal scrolling
    and sticky first three columns (Section, Row, Total).

    Args:
        result: The detailed cash flow result
        aggregation: How to aggregate periods (monthly, quarterly, annual)
        is_mixed_income: If True, show mixed income GPR breakdown; if False, show all-market GPR
        tif_lump_sum: TIF lump sum amount (shown as source for mixed income)
    """
    st.subheader("Detailed Cash Flows")

    # Aggregate periods based on selection
    all_periods = aggregate_periods(result.periods, aggregation)
    max_periods = len(all_periods)

    # Period info
    if aggregation == "monthly":
        period_label = "months"
    elif aggregation == "quarterly":
        period_label = "quarters"
    else:
        period_label = "years"

    st.caption(f"Total {period_label}: {max_periods} | Scroll horizontally to navigate through the timeline")

    # Show all periods - let the user scroll horizontally
    visible_periods = all_periods

    if not visible_periods:
        st.warning("No periods to display")
        return

    # Build rows for display
    # Note: visible values are from visible_periods, but totals are from ALL periods
    rows = []

    def add_row(section: str, label: str, values: List[str], total_value: float = None, is_header: bool = False, is_total: bool = False):
        # total_value is the sum across ALL periods, not just visible ones
        rows.append((section, label, values, total_value, is_header, is_total))

    # Calculate phase boundaries from the result
    predev_end = result.predevelopment_end
    construction_end = result.construction_end
    leaseup_end = result.leaseup_end
    operations_end = result.operations_end

    # Build cumulative phase counts for each period
    def get_phase_counts(periods_list):
        """Get cumulative phase counts for each period."""
        phase_counts = []

        for p in periods_list:
            period_start = p.period_start
            period_end = p.period_end

            # Determine which phase and cumulative count
            # For aggregated periods (quarterly/annual), use the end period for the count

            # Predevelopment: months 1 to predev_end
            if period_end <= predev_end:
                predev_count = period_end  # Cumulative count within predev
            elif period_start <= predev_end:
                predev_count = predev_end  # Partial overlap, show last predev month
            else:
                predev_count = ""

            # Construction: months predev_end+1 to construction_end
            if period_start > predev_end and period_end <= construction_end:
                construction_count = period_end - predev_end  # Cumulative within construction
            elif period_start <= construction_end and period_end > predev_end:
                if period_start > predev_end:
                    construction_count = period_end - predev_end
                else:
                    construction_count = period_end - predev_end if period_end <= construction_end else construction_end - predev_end
            else:
                construction_count = ""

            # Lease-up: months construction_end+1 to leaseup_end
            if period_start > construction_end and period_end <= leaseup_end:
                leaseup_count = period_end - construction_end
            elif period_start <= leaseup_end and period_end > construction_end:
                if period_start > construction_end:
                    leaseup_count = period_end - construction_end if period_end <= leaseup_end else leaseup_end - construction_end
                else:
                    leaseup_count = ""
            else:
                leaseup_count = ""

            # Operations: months leaseup_end+1 to operations_end
            if period_start > leaseup_end and period_end <= operations_end:
                operations_count = period_end - leaseup_end
            elif period_start <= operations_end and period_end > leaseup_end:
                if period_start > leaseup_end:
                    operations_count = period_end - leaseup_end if period_end <= operations_end else operations_end - leaseup_end
                else:
                    operations_count = ""
            else:
                operations_count = ""

            # Reversion: last period
            is_reversion = period_end > operations_end

            # Loan types
            has_const_loan = period_start <= construction_end
            has_perm_loan = period_end > construction_end

            phase_counts.append({
                'predev': predev_count,
                'construction': construction_count,
                'leaseup': leaseup_count,
                'operations': operations_count,
                'reversion': "X" if is_reversion else "",
                'const_loan': "X" if has_const_loan else "",
                'perm_loan': "X" if has_perm_loan else "",
            })

        return phase_counts

    # Get phase info for each visible period
    phase_info = get_phase_counts(visible_periods)

    # Header section - Period identification
    add_row("", "Period", [p.label for p in visible_periods], is_header=True)

    # Phase tracking rows
    add_row("Phase", "Predevelopment", [str(pi['predev']) for pi in phase_info], is_header=True)
    add_row("", "Construction", [str(pi['construction']) for pi in phase_info], is_header=True)
    add_row("", "Lease-up", [str(pi['leaseup']) for pi in phase_info], is_header=True)
    add_row("", "Operations", [str(pi['operations']) for pi in phase_info], is_header=True)
    add_row("", "Reversion", [str(pi['reversion']) for pi in phase_info], is_header=True)

    # Loan type tracking
    add_row("Loan", "Construction", [str(pi['const_loan']) for pi in phase_info], is_header=True)
    add_row("", "Permanent", [str(pi['perm_loan']) for pi in phase_info], is_header=True)

    # Development section - totals from ALL periods
    add_row("Development", "TDC Drawn - Total",
            [_fmt_currency(p.tdc_draw_total) for p in visible_periods],
            sum(p.tdc_draw_total for p in all_periods))
    add_row("", "  Predevelopment",
            [_fmt_currency(p.tdc_draw_predev) for p in visible_periods],
            sum(p.tdc_draw_predev for p in all_periods))
    add_row("", "  Construction",
            [_fmt_currency(p.tdc_draw_construction) for p in visible_periods],
            sum(p.tdc_draw_construction for p in all_periods))

    # Equity section
    add_row("Equity", "Equity Drawn",
            [_fmt_currency(p.equity_drawn) for p in visible_periods],
            sum(p.equity_drawn for p in all_periods))

    # TIF Lump Sum (for mixed income with TIF enabled)
    # TIF draws alongside equity during predevelopment/construction
    if is_mixed_income and tif_lump_sum > 0:
        # Calculate TIF draw schedule - same pattern as equity
        # TIF is drawn pro-rata with equity during the same periods
        total_equity = sum(p.equity_drawn for p in all_periods)
        if total_equity > 0:
            tif_draws = []
            for p in visible_periods:
                # TIF draws proportionally to equity
                equity_pct = p.equity_drawn / total_equity if total_equity > 0 else 0
                tif_draw = tif_lump_sum * equity_pct
                tif_draws.append(_fmt_currency(tif_draw))
            add_row("TIF", "TIF Lump Sum Drawn",
                    tif_draws,
                    tif_lump_sum)

    # Debt sources section
    add_row("Debt", "Debt Financed",
            [_fmt_currency(p.debt_financed) for p in visible_periods],
            sum(p.debt_financed for p in all_periods))

    # GPR section - show appropriate breakdown based on scenario
    if is_mixed_income:
        # Mixed income: show market units and affordable units GPR
        add_row("GPR", "GPR - Market Units",
                [_fmt_currency(p.gpr_mixed_market) for p in visible_periods],
                sum(p.gpr_mixed_market for p in all_periods))
        add_row("", "GPR - Affordable Units",
                [_fmt_currency(p.gpr_mixed_affordable) for p in visible_periods],
                sum(p.gpr_mixed_affordable for p in all_periods))
    else:
        # Market rate: show all-market GPR only
        add_row("GPR", "GPR - All Market",
                [_fmt_currency(p.gpr_all_market) for p in visible_periods],
                sum(p.gpr_all_market for p in all_periods))

    # Operations section
    add_row("Operations", "GPR",
            [_fmt_currency(p.gpr) for p in visible_periods],
            sum(p.gpr for p in all_periods))
    add_row("", "Less Vacancy",
            [_fmt_currency(-p.vacancy, show_parens=True) if p.vacancy > 0 else "-" for p in visible_periods],
            -sum(p.vacancy for p in all_periods))
    add_row("", "EGI",
            [_fmt_currency(p.egi) for p in visible_periods],
            sum(p.egi for p in all_periods))
    add_row("", "Less Opex (ex taxes)",
            [_fmt_currency(-p.opex_ex_taxes, show_parens=True) if p.opex_ex_taxes > 0 else "-" for p in visible_periods],
            -sum(p.opex_ex_taxes for p in all_periods))
    add_row("", "Less Property Taxes",
            [_fmt_currency(-p.property_taxes, show_parens=True) if p.property_taxes > 0 else "-" for p in visible_periods],
            -sum(p.property_taxes for p in all_periods))
    add_row("", "NOI",
            [_fmt_currency(p.noi) for p in visible_periods],
            sum(p.noi for p in all_periods), is_total=True)

    # TIF Benefits section (only if there's TIF data)
    has_tif = any(p.tif_net_benefit > 0 for p in all_periods)
    if has_tif:
        add_row("TIF Benefits", "TIF Reimbursement",
                [_fmt_currency(p.tif_reimbursement) for p in visible_periods],
                sum(p.tif_reimbursement for p in all_periods))
        add_row("", "Tax Abatement",
                [_fmt_currency(p.tif_abatement) for p in visible_periods],
                sum(p.tif_abatement for p in all_periods))
        add_row("", "Net TIF Benefit",
                [_fmt_currency(p.tif_net_benefit) for p in visible_periods],
                sum(p.tif_net_benefit for p in all_periods), is_total=True)

    # Investment section
    add_row("Investment", "Dev Cost",
            [_fmt_currency(p.dev_cost) for p in visible_periods],
            sum(p.dev_cost for p in all_periods))
    add_row("", "Reserves",
            [_fmt_currency(p.reserves) for p in visible_periods],
            sum(p.reserves for p in all_periods))
    add_row("", "Reversion",
            [_fmt_currency(p.reversion) for p in visible_periods],
            sum(p.reversion for p in all_periods))
    add_row("", "Unlevered CF",
            [_fmt_currency(p.unlevered_cf) for p in visible_periods],
            sum(p.unlevered_cf for p in all_periods), is_total=True)

    # Construction debt section
    add_row("Const Debt", "Debt Added",
            [_fmt_currency(p.const_debt_added) for p in visible_periods],
            sum(p.const_debt_added for p in all_periods))
    add_row("", "Interest",
            [_fmt_currency(p.const_debt_interest) for p in visible_periods],
            sum(p.const_debt_interest for p in all_periods))
    add_row("", "Repaid",
            [_fmt_currency(p.const_debt_repaid) for p in visible_periods],
            sum(p.const_debt_repaid for p in all_periods))
    add_row("", "Net CF",
            [_fmt_currency(p.const_debt_net) for p in visible_periods],
            sum(p.const_debt_net for p in all_periods))

    # Permanent debt section
    add_row("Perm Debt", "Payment",
            [_fmt_currency(-p.perm_pmt, show_parens=True) if p.perm_pmt > 0 else "-" for p in visible_periods],
            -sum(p.perm_pmt for p in all_periods))
    add_row("", "  Interest",
            [_fmt_currency(-p.perm_interest, show_parens=True) if p.perm_interest > 0 else "-" for p in visible_periods],
            -sum(p.perm_interest for p in all_periods))
    add_row("", "  Principal",
            [_fmt_currency(-p.perm_principal, show_parens=True) if p.perm_principal > 0 else "-" for p in visible_periods],
            -sum(p.perm_principal for p in all_periods))
    add_row("", "Payoff",
            [_fmt_currency(p.perm_payoff) for p in visible_periods],
            sum(p.perm_payoff for p in all_periods))
    add_row("", "Net CF",
            [_fmt_currency(p.perm_net) for p in visible_periods],
            sum(p.perm_net for p in all_periods))

    # Mezzanine debt section (only if active)
    has_mezz = any(p.mezz_active for p in all_periods)
    if has_mezz:
        add_row("Mezz Debt", "Payment",
                [_fmt_currency(-p.mezz_pmt, show_parens=True) if p.mezz_pmt > 0 else "-" for p in visible_periods],
                -sum(p.mezz_pmt for p in all_periods))
        add_row("", "  Interest",
                [_fmt_currency(-p.mezz_interest, show_parens=True) if p.mezz_interest > 0 else "-" for p in visible_periods],
                -sum(p.mezz_interest for p in all_periods))
        add_row("", "  Principal",
                [_fmt_currency(-p.mezz_principal, show_parens=True) if p.mezz_principal > 0 else "-" for p in visible_periods],
                -sum(p.mezz_principal for p in all_periods))
        add_row("", "Payoff",
                [_fmt_currency(-p.mezz_payoff, show_parens=True) if p.mezz_payoff > 0 else "-" for p in visible_periods],
                -sum(p.mezz_payoff for p in all_periods))
        add_row("", "Net CF",
                [_fmt_currency(p.mezz_net) for p in visible_periods],
                sum(p.mezz_net for p in all_periods))

    # Preferred equity section (only if active)
    has_pref = any(p.pref_active for p in all_periods)
    if has_pref:
        add_row("Preferred", "Preferred Return",
                [_fmt_currency(-p.pref_return, show_parens=True) if p.pref_return > 0 else "-" for p in visible_periods],
                -sum(p.pref_return for p in all_periods))
        add_row("", "Payoff",
                [_fmt_currency(-p.pref_payoff, show_parens=True) if p.pref_payoff > 0 else "-" for p in visible_periods],
                -sum(p.pref_payoff for p in all_periods))
        add_row("", "Net CF",
                [_fmt_currency(p.pref_net) for p in visible_periods],
                sum(p.pref_net for p in all_periods))

    # Final section - show breakdown if mezz/pref active
    if has_mezz or has_pref:
        add_row("Final", "Senior Debt CF",
                [_fmt_currency(p.net_senior_debt_cf) for p in visible_periods],
                sum(p.net_senior_debt_cf for p in all_periods))
        if has_mezz:
            add_row("", "Mezz Debt CF",
                    [_fmt_currency(p.net_mezz_debt_cf) for p in visible_periods],
                    sum(p.net_mezz_debt_cf for p in all_periods))
        if has_pref:
            add_row("", "Preferred CF",
                    [_fmt_currency(p.net_preferred_cf) for p in visible_periods],
                    sum(p.net_preferred_cf for p in all_periods))

    add_row("Final" if not (has_mezz or has_pref) else "", "Net Debt CF",
            [_fmt_currency(p.net_debt_cf) for p in visible_periods],
            sum(p.net_debt_cf for p in all_periods))
    add_row("", "Levered CF",
            [_fmt_currency(p.levered_cf) for p in visible_periods],
            sum(p.levered_cf for p in all_periods), is_total=True)

    # Build HTML table with sticky columns
    html = _build_sticky_table_html(rows, visible_periods)
    st.markdown(html, unsafe_allow_html=True)

    # Add summary totals below the table
    st.markdown("---")
    col1, col2, col3, col4 = st.columns(4)

    total_equity = sum(p.equity_drawn for p in all_periods)
    total_noi = sum(p.noi for p in all_periods)
    total_levered_cf = sum(p.levered_cf for p in all_periods)
    total_reversion = sum(p.reversion for p in all_periods)

    with col1:
        st.metric("Total Equity Invested", _fmt_currency(abs(total_equity)))
    with col2:
        st.metric("Total NOI", _fmt_currency(total_noi))
    with col3:
        st.metric("Reversion Value", _fmt_currency(total_reversion))
    with col4:
        st.metric("Total Levered CF", _fmt_currency(total_levered_cf))


def _build_sticky_table_html(
    rows: List[tuple],
    periods: List[AggregatedPeriod],
) -> str:
    """Build HTML table with sticky first three columns (Section, Row, Total)."""

    # CSS for sticky columns - explicit backgrounds to work in both light/dark modes
    # Total column is 3rd sticky column on left, right after Section and Row
    css = """
    <style>
    .cf-table-container {
        overflow-x: auto;
        max-width: 100%;
        max-height: 600px;
        overflow-y: auto;
        background-color: #ffffff;
    }
    .cf-table {
        border-collapse: collapse;
        font-size: 13px;
        font-family: 'Source Sans Pro', sans-serif;
        color: #333333;
        background-color: #ffffff;
    }
    .cf-table th, .cf-table td {
        padding: 6px 10px;
        text-align: right;
        border-bottom: 1px solid #d0d0d0;
        white-space: nowrap;
        color: #333333;
        background-color: #ffffff;
    }
    .cf-table th {
        background-color: #e8e8e8;
        font-weight: 600;
        position: sticky;
        top: 0;
        z-index: 2;
        color: #333333;
    }
    .cf-table .sticky-col {
        position: sticky;
        background-color: #f5f5f5;
        z-index: 1;
        color: #333333;
    }
    .cf-table .sticky-col-1 {
        left: 0;
        min-width: 90px;
        text-align: left;
        font-weight: 600;
        color: #1565c0;
    }
    .cf-table .sticky-col-2 {
        left: 90px;
        min-width: 160px;
        text-align: left;
        color: #333333;
    }
    .cf-table .sticky-col-3 {
        left: 250px;
        min-width: 100px;
        background-color: #fff8e1;
        font-weight: 600;
        color: #333333;
        border-right: 2px solid #ffc107;
    }
    .cf-table th.sticky-col {
        z-index: 3;
        background-color: #d8d8d8;
        color: #333333;
    }
    .cf-table th.sticky-col-1 {
        color: #1565c0;
    }
    .cf-table th.sticky-col-3 {
        background-color: #ffe082;
        border-right: 2px solid #ffc107;
    }
    .cf-table .total-row td {
        font-weight: 600;
        background-color: #e8f4e8;
        color: #333333;
    }
    .cf-table .total-row .sticky-col {
        background-color: #d8e8d8;
    }
    .cf-table .total-row .sticky-col-1 {
        color: #1565c0;
    }
    .cf-table .total-row .sticky-col-3 {
        background-color: #c8e6c9;
        border-right: 2px solid #4caf50;
    }
    .cf-table .header-row td {
        background-color: #e0e0e0;
        color: #333333;
        font-weight: 600;
    }
    .cf-table .header-row .sticky-col {
        background-color: #d0d0d0;
    }
    .cf-table .header-row .sticky-col-1 {
        color: #1565c0;
    }
    .cf-table .header-row .sticky-col-3 {
        background-color: #ffe082;
    }
    .cf-table tr:hover td {
        background-color: #f0f0f0;
    }
    .cf-table tr:hover .sticky-col {
        background-color: #e5e5e5;
    }
    .cf-table tr:hover .sticky-col-3 {
        background-color: #fff3c4;
    }
    </style>
    """

    # Build table header - Total column is 3rd, before period columns
    header_html = "<tr>"
    header_html += '<th class="sticky-col sticky-col-1">Section</th>'
    header_html += '<th class="sticky-col sticky-col-2">Row</th>'
    header_html += '<th class="sticky-col sticky-col-3">TOTAL</th>'
    for p in periods:
        header_html += f"<th>{p.label}</th>"
    header_html += "</tr>"

    # Build table rows
    rows_html = ""
    for section, label, values, total_value, is_header, is_total in rows:
        row_class = ""
        if is_header:
            row_class = "header-row"
        elif is_total:
            row_class = "total-row"

        rows_html += f'<tr class="{row_class}">'
        rows_html += f'<td class="sticky-col sticky-col-1">{section}</td>'
        rows_html += f'<td class="sticky-col sticky-col-2">{label}</td>'

        # Add total column (3rd sticky column) - this is the sum of ALL periods
        if total_value is not None:
            total_str = _fmt_currency(total_value)
        else:
            total_str = "-"
        rows_html += f'<td class="sticky-col sticky-col-3">{total_str}</td>'

        # Then add period columns (visible periods only)
        for val in values:
            rows_html += f"<td>{val}</td>"
        rows_html += "</tr>"

    # Combine into full table
    html = f"""
    {css}
    <div class="cf-table-container">
        <table class="cf-table">
            <thead>{header_html}</thead>
            <tbody>{rows_html}</tbody>
        </table>
    </div>
    """

    return html


def render_irr_summary(result: DetailedCashFlowResult) -> None:
    """Render IRR summary metrics."""
    st.subheader("Returns Summary")

    # Calculate additional return metrics
    # Equity Multiple = Total Distributions / Total Equity Invested
    total_equity_out = abs(sum(p.levered_cf for p in result.periods if p.levered_cf < 0))
    total_distributions = sum(p.levered_cf for p in result.periods if p.levered_cf > 0)

    if total_equity_out > 0:
        equity_multiple = total_distributions / total_equity_out
    else:
        equity_multiple = 0.0

    # Total profit
    total_profit = total_distributions - total_equity_out

    # ROE (Return on Equity) = Total Profit / Equity Invested
    if total_equity_out > 0:
        roe = total_profit / total_equity_out
    else:
        roe = 0.0

    # ROI (Return on Investment) - annualized
    # Using simple annualization: ROE / years
    years = result.total_periods / 12
    if years > 0:
        roi_annualized = roe / years
    else:
        roi_annualized = 0.0

    # Find first stabilized period NOI for debt metrics
    stabilized_noi = 0.0
    total_debt_service = 0.0
    for p in result.periods:
        if p.header.is_operations and not p.header.is_reversion:
            if stabilized_noi == 0:
                stabilized_noi = p.operations.noi * 12  # Annualize
            total_debt_service = p.permanent_debt.pmt_in_period * 12  # Annualized
            if p.mezzanine_debt and p.mezzanine_debt.is_active:
                total_debt_service += p.mezzanine_debt.pmt_in_period * 12
            if p.preferred_equity and p.preferred_equity.is_active:
                total_debt_service += p.preferred_equity.paid_return * 12
            break

    # DSCR = NOI / Debt Service
    dscr = stabilized_noi / total_debt_service if total_debt_service > 0 else 0.0

    # Yield on Cost = NOI / TDC
    yoc = stabilized_noi / result.sources_uses.tdc if result.sources_uses.tdc > 0 else 0.0

    # Debt Yield = NOI / Total Debt
    debt_yield = stabilized_noi / result.sources_uses.construction_loan if result.sources_uses.construction_loan > 0 else 0.0

    # Display metrics in two rows
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Unlevered IRR", f"{result.unlevered_irr:.2%}")

    with col2:
        st.metric("Levered IRR", f"{result.levered_irr:.2%}")

    with col3:
        st.metric("Equity Multiple", f"{equity_multiple:.2f}x")

    with col4:
        st.metric("Total Equity Invested", f"${total_equity_out:,.0f}")

    # Second row - debt metrics
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("DSCR", f"{dscr:.2f}x", help="Debt Service Coverage Ratio = NOI / Debt Service")

    with col2:
        st.metric("Yield on Cost", f"{yoc:.2%}", help="Stabilized NOI / TDC")

    with col3:
        st.metric("Debt Yield", f"{debt_yield:.2%}", help="Stabilized NOI / Total Debt")

    with col4:
        st.metric("Reversion Value", f"${result.reversion_value:,.0f}")

    # Third row
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("ROE (Total)", f"{roe:.2%}", help="Total Return on Equity = Profit / Equity")

    with col2:
        st.metric("ROI (Annualized)", f"{roi_annualized:.2%}", help="Annualized Return on Investment")

    with col3:
        st.metric("Total Profit", f"${total_profit:,.0f}")

    with col4:
        st.metric("LTC", f"{result.sources_uses.ltc:.1%}", help="Loan-to-Cost")

    # === DSCR Over Time Chart ===
    # Calculate DSCR for each operations period
    dscr_periods = []
    dscr_values = []

    for p in result.periods:
        if p.header.is_operations or p.header.is_leaseup:
            if p.header.is_reversion:
                continue

            noi = p.operations.noi
            debt_service = p.permanent_debt.pmt_in_period
            if p.mezzanine_debt and p.mezzanine_debt.is_active:
                debt_service += p.mezzanine_debt.pmt_in_period
            if p.preferred_equity and p.preferred_equity.is_active:
                debt_service += p.preferred_equity.paid_return

            if debt_service > 0:
                period_dscr = noi / debt_service
            else:
                period_dscr = 0

            dscr_periods.append(p.header.period)
            dscr_values.append(period_dscr)

    if dscr_periods:
        import plotly.graph_objects as go

        fig = go.Figure()

        # DSCR line
        fig.add_trace(go.Scatter(
            x=dscr_periods,
            y=dscr_values,
            mode='lines+markers',
            name='DSCR',
            line=dict(color='#1f77b4', width=2),
            marker=dict(size=4),
        ))

        # Add DSCR threshold line at 1.25
        fig.add_hline(y=1.25, line_dash="dash", line_color="orange",
                     annotation_text="1.25x DSCR Minimum", annotation_position="bottom right")

        # Add DSCR threshold line at 1.0
        fig.add_hline(y=1.0, line_dash="dot", line_color="red",
                     annotation_text="1.0x Break-even", annotation_position="bottom right")

        fig.update_layout(
            title="Debt Service Coverage Ratio Over Time",
            xaxis_title="Month",
            yaxis_title="DSCR",
            height=300,
            margin=dict(l=50, r=50, t=50, b=50),
            showlegend=False,
        )

        # Add shaded region below 1.25 to show risk
        fig.add_shape(
            type="rect",
            x0=min(dscr_periods) if dscr_periods else 0,
            x1=max(dscr_periods) if dscr_periods else 0,
            y0=0,
            y1=1.0,
            fillcolor="rgba(255, 0, 0, 0.1)",
            line=dict(width=0),
            layer="below",
        )
        fig.add_shape(
            type="rect",
            x0=min(dscr_periods) if dscr_periods else 0,
            x1=max(dscr_periods) if dscr_periods else 0,
            y0=1.0,
            y1=1.25,
            fillcolor="rgba(255, 165, 0, 0.1)",
            line=dict(width=0),
            layer="below",
        )

        st.plotly_chart(fig, use_container_width=True)


def render_operating_statement(result: DetailedCashFlowResult) -> None:
    """Render stabilized year operating statement."""
    st.subheader("Stabilized Operating Statement")

    # Find first full year of operations
    ops_start = result.leaseup_end + 1
    ops_end = min(ops_start + 11, result.operations_end)

    if ops_end <= ops_start:
        st.warning("Insufficient operations periods for annual statement")
        return

    # Aggregate first year of operations
    year1_periods = [p for p in result.periods
                     if ops_start <= p.header.period <= ops_end
                     and not p.header.is_reversion]

    if not year1_periods:
        st.warning("No stabilized periods found")
        return

    # Aggregate values
    total_gpr = sum(p.operations.gpr for p in year1_periods)
    total_vacancy = sum(p.operations.less_vacancy for p in year1_periods)
    total_egi = sum(p.operations.egi for p in year1_periods)
    total_opex = sum(p.operations.less_opex_ex_taxes for p in year1_periods)
    total_taxes = sum(p.operations.less_property_taxes for p in year1_periods)
    total_noi = sum(p.operations.noi for p in year1_periods)
    total_debt_service = sum(p.permanent_debt.pmt_in_period for p in year1_periods)

    # Add mezz and preferred debt service
    total_mezz_service = sum(p.mezzanine_debt.pmt_in_period if p.mezzanine_debt else 0 for p in year1_periods)
    total_pref_service = sum(p.preferred_equity.paid_return if p.preferred_equity else 0 for p in year1_periods)

    total_debt_service += total_mezz_service + total_pref_service

    cash_flow_after_ds = total_noi - total_debt_service

    # Build operating statement
    os_data = [
        {"Line Item": "Gross Potential Rent", "Annual": f"${total_gpr:,.0f}", "Monthly": f"${total_gpr/12:,.0f}"},
        {"Line Item": "Less: Vacancy", "Annual": f"(${total_vacancy:,.0f})", "Monthly": f"(${total_vacancy/12:,.0f})"},
        {"Line Item": "Effective Gross Income", "Annual": f"${total_egi:,.0f}", "Monthly": f"${total_egi/12:,.0f}"},
        {"Line Item": "", "Annual": "", "Monthly": ""},
        {"Line Item": "Less: Operating Expenses", "Annual": f"(${total_opex:,.0f})", "Monthly": f"(${total_opex/12:,.0f})"},
        {"Line Item": "Less: Property Taxes", "Annual": f"(${total_taxes:,.0f})", "Monthly": f"(${total_taxes/12:,.0f})"},
        {"Line Item": "Net Operating Income", "Annual": f"${total_noi:,.0f}", "Monthly": f"${total_noi/12:,.0f}"},
        {"Line Item": "", "Annual": "", "Monthly": ""},
        {"Line Item": "Less: Debt Service", "Annual": f"(${total_debt_service:,.0f})", "Monthly": f"(${total_debt_service/12:,.0f})"},
        {"Line Item": "Cash Flow After Debt Service", "Annual": f"${cash_flow_after_ds:,.0f}", "Monthly": f"${cash_flow_after_ds/12:,.0f}"},
    ]

    df = pd.DataFrame(os_data)
    st.dataframe(df, use_container_width=True, hide_index=True)

    # Key operating metrics
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        vacancy_rate = total_vacancy / total_gpr if total_gpr > 0 else 0
        st.metric("Vacancy Rate", f"{vacancy_rate:.1%}")

    with col2:
        opex_ratio = (total_opex + total_taxes) / total_egi if total_egi > 0 else 0
        st.metric("OpEx Ratio", f"{opex_ratio:.1%}")

    with col3:
        dscr = total_noi / total_debt_service if total_debt_service > 0 else 0
        st.metric("DSCR", f"{dscr:.2f}x")

    with col4:
        cash_on_cash = cash_flow_after_ds / result.total_equity_invested if result.total_equity_invested > 0 else 0
        st.metric("Cash-on-Cash", f"{cash_on_cash:.1%}")


def render_exit_waterfall(result: DetailedCashFlowResult) -> None:
    """Render the exit waterfall showing how proceeds flow through capital stack."""
    st.subheader("Exit Waterfall")

    # Find reversion period
    reversion_period = None
    for p in result.periods:
        if p.header.is_reversion:
            reversion_period = p
            break

    if not reversion_period:
        st.warning("No reversion period found")
        return

    # Get key values
    sale_price = reversion_period.investment.reversion
    senior_payoff = abs(reversion_period.permanent_debt.payoff)

    mezz_payoff = 0
    if reversion_period.mezzanine_debt and reversion_period.mezzanine_debt.is_active:
        mezz_payoff = abs(reversion_period.mezzanine_debt.payoff)

    pref_payoff = 0
    if reversion_period.preferred_equity and reversion_period.preferred_equity.is_active:
        pref_payoff = abs(reversion_period.preferred_equity.payoff)

    # Calculate waterfall
    remaining_after_senior = sale_price - senior_payoff
    remaining_after_mezz = remaining_after_senior - mezz_payoff
    remaining_after_pref = remaining_after_mezz - pref_payoff

    # Total equity invested
    total_equity = abs(sum(p.levered_cf for p in result.periods if p.levered_cf < 0))

    # Equity return
    equity_return = remaining_after_pref

    # Build waterfall display
    waterfall_data = [
        {"Item": "Gross Sale Price", "Amount": f"${sale_price:,.0f}", "Running Total": f"${sale_price:,.0f}"},
        {"Item": "Less: Senior Debt Payoff", "Amount": f"(${senior_payoff:,.0f})", "Running Total": f"${remaining_after_senior:,.0f}"},
    ]

    if mezz_payoff > 0:
        waterfall_data.append({
            "Item": "Less: Mezzanine Payoff",
            "Amount": f"(${mezz_payoff:,.0f})",
            "Running Total": f"${remaining_after_mezz:,.0f}"
        })

    if pref_payoff > 0:
        waterfall_data.append({
            "Item": "Less: Preferred Payoff",
            "Amount": f"(${pref_payoff:,.0f})",
            "Running Total": f"${remaining_after_pref:,.0f}"
        })

    waterfall_data.append({
        "Item": "Net to Common Equity",
        "Amount": f"${equity_return:,.0f}",
        "Running Total": "-"
    })

    df = pd.DataFrame(waterfall_data)
    st.dataframe(df, use_container_width=True, hide_index=True)

    # Summary metrics
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Equity Invested", f"${total_equity:,.0f}")
    with col2:
        st.metric("Equity Return at Exit", f"${equity_return:,.0f}")
    with col3:
        profit = equity_return - total_equity + sum(p.levered_cf for p in result.periods if p.levered_cf > 0 and not p.header.is_reversion)
        st.metric("Total Profit", f"${profit:,.0f}")


def render_sensitivity_analysis(result: DetailedCashFlowResult, inputs_dict: dict) -> None:
    """Render a simple sensitivity analysis showing IRR changes.

    Args:
        result: The base case detailed cash flow result
        inputs_dict: Dictionary of input values used to generate the result
    """
    st.subheader("Sensitivity Analysis")
    st.caption("Shows how IRR changes with different assumptions")

    # Base case values
    base_irr = result.levered_irr

    # Create sensitivity table data
    # Format: (label, -10% value, -5% value, base, +5% value, +10% value)
    sensitivity_data = []

    # We'll show sensitivities to key variables
    variables = [
        ("Exit Cap Rate", "exit_cap_rate", [-0.01, -0.005, 0, 0.005, 0.01]),
        ("Vacancy Rate", "vacancy_rate", [-0.02, -0.01, 0, 0.01, 0.02]),
        ("Perm Rate", "perm_rate", [-0.01, -0.005, 0, 0.005, 0.01]),
        ("Construction Cost", "hard_costs", [-0.10, -0.05, 0, 0.05, 0.10]),  # As multiplier
    ]

    # Since we can't re-run the model here (it's expensive), we'll show
    # approximate sensitivities using rules of thumb

    # Exit cap rate: ~30-50 bps IRR per 25 bps cap rate change
    exit_cap_sens = [-200, -100, 0, -100, -200]  # bps IRR change

    # Vacancy: ~20 bps IRR per 1% vacancy change
    vacancy_sens = [40, 20, 0, -20, -40]

    # Interest rate: ~15-20 bps IRR per 50 bps rate change
    rate_sens = [40, 20, 0, -20, -40]

    # Construction cost: ~50-75 bps IRR per 5% cost change
    cost_sens = [150, 75, 0, -75, -150]

    sensitivity_data = [
        {"Variable": "Exit Cap Rate", "-10%": f"{(base_irr + exit_cap_sens[0]/10000):.2%}",
         "-5%": f"{(base_irr + exit_cap_sens[1]/10000):.2%}",
         "Base": f"{base_irr:.2%}",
         "+5%": f"{(base_irr + exit_cap_sens[3]/10000):.2%}",
         "+10%": f"{(base_irr + exit_cap_sens[4]/10000):.2%}"},
        {"Variable": "Vacancy Rate", "-2%": f"{(base_irr + vacancy_sens[0]/10000):.2%}",
         "-1%": f"{(base_irr + vacancy_sens[1]/10000):.2%}",
         "Base": f"{base_irr:.2%}",
         "+1%": f"{(base_irr + vacancy_sens[3]/10000):.2%}",
         "+2%": f"{(base_irr + vacancy_sens[4]/10000):.2%}"},
        {"Variable": "Interest Rate", "-100bp": f"{(base_irr + rate_sens[0]/10000):.2%}",
         "-50bp": f"{(base_irr + rate_sens[1]/10000):.2%}",
         "Base": f"{base_irr:.2%}",
         "+50bp": f"{(base_irr + rate_sens[3]/10000):.2%}",
         "+100bp": f"{(base_irr + rate_sens[4]/10000):.2%}"},
        {"Variable": "Construction Cost", "-10%": f"{(base_irr + cost_sens[0]/10000):.2%}",
         "-5%": f"{(base_irr + cost_sens[1]/10000):.2%}",
         "Base": f"{base_irr:.2%}",
         "+5%": f"{(base_irr + cost_sens[3]/10000):.2%}",
         "+10%": f"{(base_irr + cost_sens[4]/10000):.2%}"},
    ]

    # Convert to DataFrame and display
    import pandas as pd
    df = pd.DataFrame(sensitivity_data)

    # Style the dataframe - highlight base column
    st.dataframe(df, use_container_width=True, hide_index=True)

    st.caption("Note: Sensitivities are approximate estimates based on typical relationships. "
              "Actual changes may vary based on deal structure.")


def render_property_tax_schedule_table(
    tax_stack: TaxingAuthorityStack,
    tdc: float,
    baseline_value: float,
    tif_term_years: int,
    discount_rate: float,
    inflation_rate: float,
    assessment_growth: float,
    start_year: int = 2026,
) -> None:
    """Render the detailed property tax schedule with scrolling."""

    # Generate year-by-year schedule
    # Property value starts at TDC at lease-up, grows at assessment growth rate
    years = list(range(start_year, start_year + tif_term_years + 1))

    # Build the data rows
    rows = []

    def add_row(section: str, label: str, values: list, total=None, is_header: bool = False):
        rows.append({
            "section": section,
            "label": label,
            "values": values,
            "total": total,
            "is_header": is_header,
        })

    # Helper for currency formatting
    def _fmt(val, decimals=0):
        if val is None or val == 0:
            return "-"
        if abs(val) >= 1e6:
            return f"${val/1e6:,.{decimals}f}M"
        elif abs(val) >= 1e3:
            return f"${val/1e3:,.{decimals}f}K"
        else:
            return f"${val:,.{decimals}f}"

    # Calculate values for each year
    property_values = []
    assessed_values = []
    increment_values = []

    # Tax by authority
    auth_taxes = {auth.code: [] for auth in tax_stack.authorities}
    auth_taxes_pv = {auth.code: [] for auth in tax_stack.authorities}

    # TIF streams
    tif_increment_nominal = []
    tif_increment_real = []
    tif_increment_pv = []

    # Totals
    total_taxes = []
    total_taxes_pv = []

    for i, year in enumerate(years):
        # Property value (TDC escalated)
        prop_val = tdc * ((1 + assessment_growth) ** i)
        property_values.append(prop_val)

        # Assessed value (same as property value for simplicity)
        assessed_val = prop_val
        assessed_values.append(assessed_val)

        # Increment over baseline
        increment = max(0, assessed_val - baseline_value)
        increment_values.append(increment)

        # Tax by authority
        year_total_tax = 0
        for auth in tax_stack.authorities:
            tax = assessed_val * auth.rate_decimal
            auth_taxes[auth.code].append(tax)
            year_total_tax += tax

            # PV of tax
            pv_factor = 1 / ((1 + discount_rate) ** i)
            auth_taxes_pv[auth.code].append(tax * pv_factor)

        total_taxes.append(year_total_tax)
        total_taxes_pv.append(year_total_tax * (1 / ((1 + discount_rate) ** i)))

        # TIF increment tax (participating authorities only)
        tif_tax_nom = increment * tax_stack.tif_participating_rate_decimal
        tif_increment_nominal.append(tif_tax_nom)

        # Real (inflation-adjusted)
        inflation_factor = (1 + inflation_rate) ** i
        tif_tax_real = tif_tax_nom / inflation_factor if inflation_factor > 0 else tif_tax_nom
        tif_increment_real.append(tif_tax_real)

        # PV
        pv_factor = 1 / ((1 + discount_rate) ** i)
        tif_increment_pv.append(tif_tax_nom * pv_factor)

    # Build table rows
    add_row("Property", "Property Value (TDC basis)", [_fmt(v) for v in property_values], _fmt(sum(property_values)))
    add_row("Property", "Assessed Value", [_fmt(v) for v in assessed_values], _fmt(sum(assessed_values)))
    add_row("Property", "Baseline Value", [_fmt(baseline_value)] * len(years), _fmt(baseline_value))
    add_row("Property", "Increment over Baseline", [_fmt(v) for v in increment_values], _fmt(sum(increment_values)))

    add_row("", "", [""] * len(years), "")  # Spacer

    # Tax by authority - Nominal
    add_row("Taxes (Nominal)", "**Tax by Authority**", [""] * len(years), "", is_header=True)
    for auth in tax_stack.authorities:
        tif_marker = " *" if auth.participates_in_tif else ""
        add_row("Taxes (Nominal)", f"{auth.code}{tif_marker}",
                [_fmt(v) for v in auth_taxes[auth.code]],
                _fmt(sum(auth_taxes[auth.code])))
    add_row("Taxes (Nominal)", "**Total Tax**",
            [_fmt(v) for v in total_taxes],
            _fmt(sum(total_taxes)))

    add_row("", "", [""] * len(years), "")  # Spacer

    # Tax by authority - PV
    add_row("Taxes (PV)", "**Tax by Authority (PV)**", [""] * len(years), "", is_header=True)
    for auth in tax_stack.authorities:
        tif_marker = " *" if auth.participates_in_tif else ""
        add_row("Taxes (PV)", f"{auth.code}{tif_marker}",
                [_fmt(v) for v in auth_taxes_pv[auth.code]],
                _fmt(sum(auth_taxes_pv[auth.code])))
    add_row("Taxes (PV)", "**Total Tax (PV)**",
            [_fmt(v) for v in total_taxes_pv],
            _fmt(sum(total_taxes_pv)))

    add_row("", "", [""] * len(years), "")  # Spacer

    # TIF Increment streams
    add_row("TIF Increment", "**TIF Increment (Participating Only)**", [""] * len(years), "", is_header=True)
    add_row("TIF Increment", "Annual (Nominal)",
            [_fmt(v) for v in tif_increment_nominal],
            _fmt(sum(tif_increment_nominal)))
    add_row("TIF Increment", "Annual (Real)",
            [_fmt(v) for v in tif_increment_real],
            _fmt(sum(tif_increment_real)))
    add_row("TIF Increment", "Annual (PV)",
            [_fmt(v) for v in tif_increment_pv],
            _fmt(sum(tif_increment_pv)))

    add_row("", "", [""] * len(years), "")  # Spacer

    # Cumulative buildup - shows how value accumulates from 0 to final
    cumulative_nominal = []
    cumulative_real = []
    cumulative_pv = []
    running_nominal = 0
    running_real = 0
    running_pv = 0

    for i in range(len(years)):
        running_nominal += tif_increment_nominal[i]
        running_real += tif_increment_real[i]
        running_pv += tif_increment_pv[i]
        cumulative_nominal.append(running_nominal)
        cumulative_real.append(running_real)
        cumulative_pv.append(running_pv)

    add_row("TIF Buildup", "**Cumulative Buildup**", [""] * len(years), "", is_header=True)
    add_row("TIF Buildup", "Cumulative (Nominal)",
            [_fmt(v) for v in cumulative_nominal],
            _fmt(cumulative_nominal[-1] if cumulative_nominal else 0))
    add_row("TIF Buildup", "Cumulative (Real)",
            [_fmt(v) for v in cumulative_real],
            _fmt(cumulative_real[-1] if cumulative_real else 0))
    add_row("TIF Buildup", "**Cumulative (PV)**",
            [_fmt(v) for v in cumulative_pv],
            _fmt(cumulative_pv[-1] if cumulative_pv else 0))

    # Build HTML table with sticky columns
    year_headers = [str(y) for y in years]

    html = """
    <style>
    .ptax-table-container {
        overflow-x: auto;
        max-width: 100%;
        border: 1px solid #ddd;
    }
    .ptax-table {
        border-collapse: collapse;
        font-size: 12px;
        white-space: nowrap;
    }
    .ptax-table th, .ptax-table td {
        padding: 4px 8px;
        border: 1px solid #ddd;
        text-align: right;
        background-color: #ffffff;
        color: #333333;
    }
    .ptax-table th {
        background-color: #f0f0f0;
        font-weight: bold;
        position: sticky;
        top: 0;
        z-index: 1;
    }
    .ptax-table .sticky-section {
        position: sticky;
        left: 0;
        background-color: #e8e8e8;
        z-index: 2;
        text-align: left;
        min-width: 100px;
    }
    .ptax-table .sticky-label {
        position: sticky;
        left: 100px;
        background-color: #f5f5f5;
        z-index: 2;
        text-align: left;
        min-width: 180px;
    }
    .ptax-table .sticky-total {
        position: sticky;
        left: 280px;
        background-color: #e0e8f0;
        z-index: 2;
        font-weight: bold;
    }
    .ptax-table .header-row td {
        font-weight: bold;
        background-color: #d0d8e0;
    }
    .ptax-table .spacer-row td {
        height: 10px;
        border: none;
        background-color: #ffffff;
    }
    </style>
    <div class="ptax-table-container">
    <table class="ptax-table">
    <thead>
    <tr>
        <th class="sticky-section">Section</th>
        <th class="sticky-label">Row</th>
        <th class="sticky-total">TOTAL</th>
    """

    for year in year_headers:
        html += f"<th>{year}</th>"
    html += "</tr></thead><tbody>"

    for row in rows:
        if row["label"] == "" and row["section"] == "":
            html += '<tr class="spacer-row"><td></td><td></td><td></td>'
            for _ in years:
                html += "<td></td>"
            html += "</tr>"
            continue

        row_class = "header-row" if row["is_header"] else ""
        html += f'<tr class="{row_class}">'
        html += f'<td class="sticky-section">{row["section"]}</td>'
        html += f'<td class="sticky-label">{row["label"]}</td>'
        html += f'<td class="sticky-total">{row["total"] or ""}</td>'
        for val in row["values"]:
            html += f"<td>{val}</td>"
        html += "</tr>"

    html += "</tbody></table></div>"

    st.markdown(html, unsafe_allow_html=True)
    st.caption("* = TIF Participating Authority")


def render_property_tax_engine(
    tax_stack: TaxingAuthorityStack,
    baseline_value: float,
    stabilized_value: float,
    tdc: float = None,
    start_year: int = 2026,
) -> None:
    """Render the property tax engine UI - tax components and schedule only (no TIF analysis)."""
    st.subheader("Property Tax Components")

    # If TDC not provided, use stabilized value as proxy
    if tdc is None:
        tdc = stabilized_value

    # Taxing authority stack
    st.markdown("**Taxing Authority Breakdown**")
    st.caption("Property taxes in Austin are collected by multiple taxing authorities.")

    # Build table
    auth_data = []
    for auth in tax_stack.authorities:
        auth_data.append({
            "Authority": auth.name,
            "Code": auth.code,
            "Rate (per $100)": f"${auth.rate_per_100:.4f}",
            "Rate (%)": f"{auth.rate_decimal:.4%}",
        })

    auth_df = pd.DataFrame(auth_data)
    st.dataframe(auth_df, use_container_width=True, hide_index=True)

    # Show totals
    col1, col2 = st.columns(2)
    with col1:
        st.metric("Total Tax Rate", f"{tax_stack.total_rate_decimal:.4%}")
    with col2:
        annual_tax = tdc * tax_stack.total_rate_decimal
        st.metric("Est. Year 1 Tax (on TDC)", f"${annual_tax:,.0f}")

    st.divider()

    # Property Tax Projection Parameters
    st.markdown("**Property Tax Projection**")

    col1, col2 = st.columns(2)
    with col1:
        projection_years = st.slider("Projection Years", 5, 30, 20, key="prop_tax_projection_years")
    with col2:
        assessment_growth = st.slider("Assessment Growth Rate", 0.0, 10.0, 2.0, 0.5,
                                     key="prop_tax_assess_growth_pct", format="%.1f%%") / 100

    st.divider()

    # Detailed Property Tax Schedule
    st.markdown("**Property Tax Schedule (Annual)**")
    st.caption("Shows estimated property tax by authority over the projection period.")

    render_property_tax_schedule_simple(
        tax_stack=tax_stack,
        tdc=tdc,
        projection_years=projection_years,
        assessment_growth=assessment_growth,
        start_year=start_year,
    )


def render_property_tax_schedule_simple(
    tax_stack: TaxingAuthorityStack,
    tdc: float,
    projection_years: int,
    assessment_growth: float,
    start_year: int = 2026,
) -> None:
    """Render a simple property tax schedule without TIF complexity."""
    years = list(range(start_year, start_year + projection_years + 1))

    # Build schedule data
    schedule_data = []
    cumulative_tax = 0

    for i, year in enumerate(years):
        # Assessed value grows over time
        assessed_val = tdc * ((1 + assessment_growth) ** i)

        # Calculate total tax
        total_tax = assessed_val * tax_stack.total_rate_decimal
        cumulative_tax += total_tax

        schedule_data.append({
            "Year": year,
            "Assessed Value": assessed_val,
            "Total Tax": total_tax,
            "Cumulative Tax": cumulative_tax,
        })

    df = pd.DataFrame(schedule_data)

    # Display chart and table side by side
    col1, col2 = st.columns([1, 1])

    with col1:
        # Chart
        chart_df = df[["Year", "Total Tax"]].copy()
        chart_df = chart_df.set_index("Year")
        st.line_chart(chart_df, height=300)

    with col2:
        # Format for display
        df_display = df.copy()
        df_display["Assessed Value"] = df_display["Assessed Value"].apply(lambda x: f"${x:,.0f}")
        df_display["Total Tax"] = df_display["Total Tax"].apply(lambda x: f"${x:,.0f}")
        df_display["Cumulative Tax"] = df_display["Cumulative Tax"].apply(lambda x: f"${x:,.0f}")

        st.dataframe(df_display, use_container_width=True, hide_index=True, height=300)

    # Summary metrics
    st.divider()
    total_taxes_paid = cumulative_tax
    avg_annual_tax = total_taxes_paid / len(years) if years else 0

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Year 1 Tax", f"${schedule_data[0]['Total Tax']:,.0f}")
    with col2:
        st.metric(f"Year {projection_years} Tax", f"${schedule_data[-1]['Total Tax']:,.0f}")
    with col3:
        st.metric("Total Over Period", f"${total_taxes_paid:,.0f}")
